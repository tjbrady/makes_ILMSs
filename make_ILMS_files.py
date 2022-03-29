#
# This part takes the csv constructed in the makeCSVs function and builds
# individual spreadsheets with formula and formatting as requested

def makeExcel(file):
    import os
    import csv
    import openpyxl
    from openpyxl.styles import Border, Side
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import Alignment

    # get name of path to the script file for opening files.
    pathTemp = os.path.dirname(__file__)
    csvFileName = pathTemp + "\\" + file
    xlsFileName = csvFileName.replace('csv', 'xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active


    with open(csvFileName) as f:
        count = 0
        reader = csv.reader(f, delimiter=',')


        for row in reader:
            ws.append(row)
            count = count + 1  # capture the total number of rows
    bottomRow = count # used to set bottom of table range
    #print(bottomRow)

    markWidth = 10 # column width for all mark columns
    ws.column_dimensions['A'].width = 20 # colA set to 20
    ws.column_dimensions['B'].width = 33 # ColB set to 33
    ws.column_dimensions['C'].width = markWidth
    ws.column_dimensions['D'].width = markWidth
    ws.column_dimensions['E'].width = markWidth
    ws.column_dimensions['F'].width = markWidth
    ws.column_dimensions['G'].width = markWidth
    ws.column_dimensions['H'].width = markWidth
    ws.column_dimensions['I'].width = markWidth
    ws.column_dimensions['J'].width = markWidth
    ws.column_dimensions['K'].width = markWidth
    ws.column_dimensions['L'].width = markWidth
    ws.column_dimensions['M'].width = markWidth
    ws.column_dimensions['N'].width = markWidth
    ws.column_dimensions['O'].width = markWidth
    ws.column_dimensions['P'].width = markWidth


    # set background colours for TL form and heading for table
    yellow = "00FFFFD9"
    ltGreen = "00EBF1DE"
    pink = "00EBD4D1"
    for rows in ws.iter_rows(min_row=7, max_row=10, min_col=1, max_col=2):
        for cell in rows:
            cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type="solid")


    ws.cell(row=12, column=3).font = Font(bold=True) # set bold font for Original Examiner
    ws.cell(row=12, column=10).font = Font(bold=True) # set bold font for Modified marks
    ws.cell(row=12, column=3).fill= PatternFill(start_color=ltGreen, end_color=ltGreen, fill_type="solid") # ltGreen for Original Examiner
    ws.cell(row=12, column=10).fill = PatternFill(start_color=pink, end_color=pink, fill_type="solid") # pink for Modified Marks
 

    ws.merge_cells("c12:I12") # Merged cells for Original Examiner
    ws.merge_cells("j12:p12") # Merged cells for Modified Marks

    # set thick black lines for TL and Assessor signoff table
    black = "00000000"
    thick = Side(border_style="thick", color=black)
    ws["A7"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    ws["A8"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    ws["A9"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    ws["A10"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    ws["B7"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    ws["B8"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    ws["B9"].border = Border(top=thick, left=thick, right=thick, bottom=thick)
    ws["B10"].border = Border(top=thick, left=thick, right=thick, bottom=thick)

    # variable to set range for creating the table
    tabRng = "A13:p" + str(bottomRow)

    tab = Table(displayName="Table1", ref=tabRng, autoFilter=None) # the autofilter here was attempt to not have filter - didnt work

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleLight19", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    # this is where the table is added
    ws.add_table(tab)

    # ws.auto_filter.add_sort_condition = False # attempts to remove filter
    # ws.auto_filter.add_sort_condition("A13", False) # attempts to remove filter
    # ws.auto_filter.ref = "A13:P13" # attempts to remove filter
    # ws.auto_filter.add_sort_condition(None) # attempts to remove filter

    ws.freeze_panes = "c14" # this freezes the top 12 rows and the left 2 columns


    # insert if formula for summing original marks unless a mod was made by TL
    rowNum = 14
    for rows in ws.iter_rows(min_row=14, max_row=bottomRow, min_col=3, max_col=3):
        for cell in rows:
            cell.value = ("=if(J" + str(rowNum) +"=0,SUM(D" + str(rowNum) + ":I" + str(rowNum) + "),\"Changed\")")
            cell.alignment = Alignment(horizontal='center')
            rowNum = rowNum +1

    # insert formula for summing modified marks unless a mod was made by TL
    rowNum = 14
    for rows in ws.iter_rows(min_row=14, max_row=bottomRow, min_col=10, max_col=10):
        for cell in rows:
            cell.value = ("=SUM(K" + str(rowNum) + ":P" + str(rowNum) + ")")
            cell.alignment = Alignment(horizontal='center')
            rowNum = rowNum +1


    # save the spreadsheet file
    wb.save(xlsFileName)
    # remove the csv file now that it is not needed
    os.remove(csvFileName)

# This function takes the pdf report from EOL as input and creates a CSL file per page
# it handles tha last page as it does not have any candidates
def makeCSVs(file):
    import re
    import pdfplumber
    import os

    # get name of path to the script file for opening files.
    pathTemp = os.path.dirname(__file__)

    # open up pdfplumber to readin the pdf file
    with pdfplumber.open(file) as pdf:
        #print(len(pdf.pages))
        #set pageNum variable to add to the filenames createdf
        pageName=""

        # pdf.pages is a list of all pages in pdf
        # so work through a page at a time.
        for page in pdf.pages:
            #capture page number a str to name output filesd
            pageName = str(page.page_number)
            print(pageName) # just to show progress in terminal window

            # convert pdf content to text
            text = page.extract_text(x_tolerance=2, y_tolerance=0)

            #initialise two variables used for making the initial csv file
            line_str = ""
            all_lines = ""

            # work through each page a pdf line at a time
            for line in text.split('\n'):
                if "CANDIDATES:" in line: # insert into csv new table and columns for ilms
                    line_str = 'TeamLeader\nTLsignoff\nAssName\nAssSignoff\n\n,,Original Examiner,,,,,,,Team Leader Marks\nCandNum,CandName,Tot,Q1mark,Q2mark,Q3mark,Q4mark,Q5mark,Q6mark,TotNew,Q1_New,Q2_New,Q3_New,Q4_New,Q5_New,Q6_New\n'
                elif "SUBJECT:" in line or "PAPER:" in line: # tidy up line with subject in it
                    line_str = line.lstrip()
                    line_str = line_str.replace(":",",") + "\n"
                elif "CENTRE:" in line: # tidy up line with centre info and capture variable centre_num for filename
                    line_str = line.lstrip()
                    line_str = line_str.replace(":",",") + "\n\n"
                    centre_num = line[10:15]
                    #print(centre_num)
                    csvName = centre_num + "_Page" + pageName +".csv" #construct filename for all the individual csvs
                    #csvName = centre_num + ".csv"  # construct filename for all the individual csvs
                # the below regex finds 4 digits so these lines are the candidate data
                elif re.search(r'\b\d\d\d\d\b', line) and "PEARSON EDEXCEL GCSE EXAMINATIONS" not in line: # find 4digit using regex
                    candNum = line[10:14] # extract cand Num
                    candName = line[16:]  # extract cand Name
                    line_str = candNum + "," + candName + "\n" # construct all the csv rows with cand data
                elif "CANDIDATE" in line: # tidy up the row with CANDIDATE LIST in it
                    line_str = line.lstrip() + "\n"
                elif "*** END OF REPORT AT:" in line: # this is the final page so should not be made into csv or xlsx
                    quit()  # quit at this point
                else: # finally split up the first row so it is easier to fit into the csv
                    line = re.sub('\s+',' ',line) # use regex to get rid of multiple speces
                    line = line.lstrip() # remove leading spaces
                    line = line.replace("PEARSON" , ",PEARSON") # insert comma
                    line = line.replace("EXAMINATIONS", "EXAMINATIONS,,") # insert comma
                    #print(line)
                    line_str = line.lstrip() +"\n"

                all_lines = all_lines + line_str


            outFile = open(pathTemp + "/" + csvName, "w")
            outFile.write(all_lines)
            outFile.close()

            makeExcel(csvName)




# Main calls the makeCSV function and pass in teh name of the PDF report file
if __name__ == '__main__':
   makeCSVs("PWP7190X.pdf")



